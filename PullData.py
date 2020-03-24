import json, requests, sys, pandas, datetime, openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import config as cfg
import constant


def format_date(p_date):
    data = datetime.datetime.strptime(p_date, '%Y-%m-%dT%H:%M:%S')
    new_date = data.date().strftime("%d/%m/%Y")
    return new_date

def format_cell_style():
    red_fill = PatternFill(start_color='EE1111',end_color = 'EE1111',fill_type = 'solid')
    orange_fill = PatternFill(start_color='f77d26',end_color='f77d26',fill_type='solid')
    yellow_fill = PatternFill(start_color='f7cd26',end_color='f7cd26',fill_type='solid')
    green_fill = PatternFill(start_color='92f726',end_color='92f726',fill_type='solid')

    for i in range(2,len(worksheet[constant.DEADLINE_COLUMN]) + 1):
        worksheet.conditional_formatting.add(constant.FORMULAE_COLUMN + str(i),FormulaRule(formula = [constant.OK_COLUMN + str(i) + '="ok"'],fill = green_fill))
        worksheet.conditional_formatting.add(constant.FORMULAE_COLUMN + str(i),CellIsRule(operator='<', formula=['TODAY()'], stopIfTrue=True, fill=red_fill))
        worksheet.conditional_formatting.add(constant.FORMULAE_COLUMN + str(i), CellIsRule(operator='=', formula=['TODAY()'], stopIfTrue=True,fill=orange_fill))
        worksheet.conditional_formatting.add(constant.FORMULAE_COLUMN + str(i), CellIsRule(operator='<=', formula=['TODAY() + 3'], stopIfTrue=True,fill=yellow_fill))


def format_title(title):
    new_title = title.replace(';', ' ')
    return new_title


def save_xlsx_file(wb):
    wb.save('output.xlsx')


def load_xlsx_file():
    return load_workbook('output.xlsx')


def create_xlsx_headers(ws):
    count = 0
    for i in range(1, 6):
        if ws.cell(row=1, column=i).value is None:
            break
        count += 1

    if count == 0:
        print('populate headers')
        ws[constant.DATE_COLUMN + '1'] = 'Data de publicação'
        ws.column_dimensions[constant.DATE_COLUMN].width = 25
        ws[constant.TITLE_COLUMN + '1'] = 'Titulo'
        ws.column_dimensions[constant.TITLE_COLUMN].width = 40
        ws[constant.TEXT_COLUMN + '1'] = 'Publicação'
        ws.column_dimensions[constant.TEXT_COLUMN].width = 40
        ws[constant.DEADLINE_COLUMN + '1'] = 'Prazo'
        ws.column_dimensions[constant.DEADLINE_COLUMN].width = 25
        ws[constant.FORMULAE_COLUMN + '1'] = 'Data de Prazo'
        ws.column_dimensions[constant.FORMULAE_COLUMN].width = 25
        ws[constant.ID_COLUMN + '1'] = 'id'


def request_url(api_key, date, diferencial):
    url = 'https://cadastroapi.aasp.org.br/api/Associado/intimacao?chave=' + api_key + '&data='+date+'&diferencial='+ diferencial
    response = requests.get(url)
    response.raise_for_status()
    return response.text


def add_formulae_to_deadline_column():
    for row_number in range(2, len(worksheet[constant.FORMULAE_COLUMN]) + 1):
        new_value = '=' + constant.DATE_COLUMN + str(row_number) + '+' + constant.DEADLINE_COLUMN + str(row_number)
        cell = worksheet.cell(row=row_number, column=constant.FORMULAE_COLUMN_NUMBER, value=new_value)
        cell.style = get_date_style()


def row_exists(row_id):
    for i in range(2, len(worksheet[constant.ID_COLUMN])+1):
        if worksheet[constant.ID_COLUMN + str(i)].value == row_id:
            print('Duplicate found id: '+ str(row_id))
            return True
    return False


def add_cells_to_worksheet(date, title, text, row_id, row_count):
    if row_exists(row_id):
        return
    date_cell = worksheet.cell(row=row_count, column=constant.DATE_COLUMN_NUMBER, value=date)
    date_cell.style = get_date_style()
    title_cell = worksheet.cell(row=row_count, column=constant.TITLE_COLUMN_NUMBER, value=title)
    text_cell = worksheet.cell(row=row_count, column=constant.TEXT_COLUMN_NUMBER, value=text[0:500])
    text_cell.alignment = Alignment(horizontal='general', vertical='top', text_rotation=0, wrap_text=True,
                                    shrink_to_fit=False, indent=0)
    id_cell = worksheet.cell(row=row_count, column=constant.ID_COLUMN_NUMBER, value=row_id)


def parse_data_to_cells():
    number_of_entries = len(worksheet['A'])
    for idx, d in enumerate(chosen_data['intimacoes']):
        row_count = idx + number_of_entries + 1
        date = format_date(d['jornal']['dataDisponibilizacao_Publicacao'])
        title = format_title(d['titulo'])
        text = format_title(d['textoPublicacao'])
        row_id = d['codigoRelacionamento']
        add_cells_to_worksheet(date, title, text, row_id, row_count)


def get_workbook():
    try:
        _workbook = load_xlsx_file()

    except FileNotFoundError:
        _workbook = Workbook()

    return _workbook


def get_date_style():
    style = NamedStyle(name='new_datetime', number_format='DD/MM/YYYY')
    try:
        workbook.add_named_style(style)
    except ValueError:
        style = 'new_datetime'
    return style


def hide_id_column():
    worksheet.column_dimensions[constant.ID_COLUMN].hidden = True


def main(date):
    global workbook
    global worksheet
    global chosen_data
    workbook = get_workbook()
    worksheet = workbook.active
    create_xlsx_headers(worksheet)

    # Main code for pulling data from AASP
    chosen_data = pandas.read_json(request_url(cfg.api_key, date, 'false'))
    parse_data_to_cells()
    add_formulae_to_deadline_column()
    hide_id_column()
    format_cell_style()
    save_xlsx_file(workbook)
